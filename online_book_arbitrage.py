#! python3
# online_book_arbitrage.py -- Uses workbook of Ebay links to open each link, get the ISBN, search the ISBN on BookScouter, and if the price on Ebay is less than the price on BookScouter, store the information to an xls workbook.

import requests, bs4, openpyxl, re, validators
from bs4 import BeautifulSoup
from requests_html import HTMLSession
from openpyxl import Workbook
from openpyxl import load_workbook

#Creates workbook. 
wb = Workbook()

#Gets sheet. 
sheet = wb.active

#loads ebay link workbook
link_workbook = load_workbook('ebay_links.xls')
link_sheet = link_workbook.active

#Writes info to workbook.
def write_to_excel(row_index, column_index, text):
	sheet.cell(row=row_index, column=column_index).value = text	

#Get page from link and convert to BeautifulSoup for html parsing.
def get_soup(url):
	res = requests.get(url)
	res.raise_for_status()
	soup = bs4.BeautifulSoup(res.text, features="lxml")
	return soup

#Returns a new row from the workbook sheet.
def get_new_row(sheet):
	new_row = sheet.max_row + 1
	return new_row

#Scrapes for ISBN.
def scrape(soup):
	if not soup.find(itemprop="productID"):
		text = 'no_product_id'
		return text
	else:
		text = soup.find(itemprop="productID").getText()
		return text

#Searches for Ebay price.
def search_ebay(url):
	session = HTMLSession()
	r = session.get(url)
	if not r.html.find('#prcIsum', first=True):
		no_price = 'no_ebay_price'
		session.close()
		return no_price
	else:
		price = r.html.find('#prcIsum', first=True).text
		price_re = re.findall("\d+\.\d+", price)
		price_decimal = price_re[0]
		session.close()
		return price_decimal	

#Searches for BookScouter price.
def search_bookscouter(isbn):
	session = HTMLSession()
	url = 'https://bookscouter.com/sell/' + isbn
	print(url)
	r = session.get(url)
	r.html.render()
	
	if not r.html.find('div.price__child.price__price.flex-child__auto', first=True):
		no_price = 'no_book_price'
		session.close()
		return no_price
	else:	
		price = r.html.find('div.price__child.price__price.flex-child__auto', first=True).text
		print(price)
		price_re = re.findall("\d+\.\d+", price)
		price_decimal = price_re[0]
		session.close()
		return price_decimal		

i = 1
write_to_excel(1, 1, 'Ebay Url')
write_to_excel(1, 2, 'Ebay Price')
write_to_excel(1, 3, 'Bookscouter Price')
write_to_excel(1, 4, 'Profit')

while i < link_sheet.max_row + 1:	

	ebay_url = link_sheet.cell(row=i, column=1).value
	print('Searching post: ' + str(i))

	if validators.url(ebay_url):
		ebay_soup = get_soup(ebay_url)
		ebay_isbn = scrape(ebay_soup)
		print(ebay_isbn)

		if ebay_isbn != 'no_product_id':
					
			bookscouter_price = search_bookscouter(ebay_isbn)
			print(bookscouter_price)

			if bookscouter_price != 'no_book_price':
				
				ebay_price = search_ebay(ebay_url)
				print(ebay_price)

				if ebay_price != 'no_ebay_price':

					if float(ebay_price) < float(bookscouter_price):
						new_row = get_new_row(sheet)
						profit = float(bookscouter_price) - float(ebay_price)
						write_to_excel(new_row, 1, ebay_url)
						write_to_excel(new_row, 2, ebay_price)
						write_to_excel(new_row, 3, bookscouter_price)
						write_to_excel(new_row, 4, str(profit))
						wb.save('book_arbitrage.xls')

	elif not validators.url(ebay_url):
		print('not valid')

	i = i + 1
