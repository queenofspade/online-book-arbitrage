#! python3
# ebay_post_link_scraper.py -- Searches eBay for used textbooks and stores links to each posting in xls workbook.

import requests, bs4, openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook

#Sets the starting url.
#url = 'https://www.ebay.com/sch/i.html?_from=R40&_sacat=0&_udlo&_udhi&LH_Time=1&_ftrt=903&_ftrv=24&_sabdlo&_sabdhi&_samilow&_samihi&_sadis=15&_stpos=31069&_sop=12&_dmd=1&_ipg=200&_fosrp=1&_nkw=used%20textbook&rt=nc&LH_FS=1&_trksid=p2045573.m1684'

#Creates workbook. 
wb = Workbook() 

#Gets sheet. 
sheet = wb.active

#Get page from link and convert to BeautifulSoup for html parsing.
def get_soup(url):
	res = requests.get(url)
	res.raise_for_status()
	soup = bs4.BeautifulSoup(res.text, features="lxml")
	return soup

#Returns a new row from the workbook sheet.
def get_new_row(sheet):
	new_row = 0
	while new_row < 65537: #65536 is the max number of rows allowed in an xls spreadsheet.
		new_row = sheet.max_row
		return new_row

#Writes info to workbook.
def write_to_excel(row_index, column_index, text):
	sheet.cell(row=row_index, column=column_index).value = text	

#Scrapes listings for post links.		
def scrape_links(element, soup, sheet, column):
	elem = soup.select(element)
	num = len(elem)		
	for i in range(num):
		text = elem[i].get('href')
		new_row = get_new_row(sheet) + 1
		write_to_excel(new_row, column, text)

#Goes to next page and gets new url.
def get_new_url(i):
	url_end = str(i)
	url = 'https://www.ebay.com/sch/i.html?_from=R40&_sacat=0&_udlo=&_udhi=&LH_Time=1&_ftrt=903&_ftrv=24&_sabdlo=&_sabdhi=&_samilow=&_samihi=&_sadis=15&_stpos=31069&_sop=12&_dmd=1&_fosrp=1&LH_FS=1&_nkw=used+textbook&_pgn=' + url_end
	print('Scraping post links for: ' + url)	
	return url

i = 1

print("How many pages would you like to scrape?")
pages = int(input())

while i < pages + 1:

	try:
		
		#Calls the function 'get_new_url' and sets this 'url' variable to its return, which is the url of the next page of listings.
		url = get_new_url(i)
		#Calls the function 'get_soup' and sets this 'soup' variable to its return, which is the page from the url converted to BeautifulSoup for html parsing.
		soup = get_soup(url)		
		#Scrapes for links using the inspected element class of the link.	
		scrape_links('a.vip', soup, sheet, 1)

		i = i + 1

	#Ends the loop when a url is invalid.	
	except requests.exceptions.RequestException:
		break	

print('Completed scraping post links. ' + str(sheet.max_row) + ' posts found.')		

#Saves the final workbook.
wb.save('ebay_links.xls')
print('Workbook saved.')